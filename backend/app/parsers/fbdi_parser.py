"""Robust Oracle FBDI template parser.

Real Oracle FBDI templates contain:
- An "Instructions and CSV Generation" sheet with overview text.
- One or more data sheets where:
  * Column 1 contains metadata labels (Name, Description, Data Type, then a row
    per applicable module e.g. "Demand Management", "Global Order Promising"...).
  * Columns 2..N contain field metadata. Field names prefixed with "*" are
    required (also indicated by a "Required" cell in the module rows).

This parser is resilient to small variations and falls back gracefully so that
manual metadata correction is always possible in the UI.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_DATA_TYPE_RE = re.compile(r"^\s*([A-Za-z]+)\s*(?:\(\s*(\d+)(?:\s*,\s*\d+)?\s*\))?", re.IGNORECASE)


def _parse_data_type(raw: str | None) -> tuple[str, int | None]:
    if not raw:
        return ("Character", None)
    m = _DATA_TYPE_RE.match(str(raw))
    if not m:
        return (str(raw).strip(), None)
    dtype = m.group(1).strip().capitalize()
    length = int(m.group(2)) if m.group(2) else None
    return (dtype, length)


def _looks_like_module_label(value: Any) -> bool:
    if not value:
        return False
    s = str(value).strip()
    keywords = (
        "management", "planning", "order promising", "operations",
        "supply", "common", "interface", "loader",
    )
    return any(k in s.lower() for k in keywords) and len(s) < 80


def parse_fbdi_template(file_path: Path | str) -> dict[str, Any]:
    """Parse an Oracle FBDI template file.

    Returns a dict with: business_object, sheets[], fields[].
    Each field has: name, display_name, description, required, data_type,
    max_length, sequence, sheet_name, sample_value, lookup_type, required_modules.
    """
    file_path = Path(file_path)
    wb = load_workbook(filename=file_path, data_only=True, keep_vba=False, read_only=False)

    business_object: str | None = None
    description: str | None = None
    # Try to extract business object name from the instructions sheet
    for sname in wb.sheetnames:
        if "instruction" in sname.lower():
            ws = wb[sname]
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and ":" in cell and len(cell) < 200:
                        if any(k in cell.lower() for k in ("upload:", "import:", "interface:")):
                            business_object = cell.split(":", 1)[1].strip()
                            break
                if business_object:
                    break
            break

    sheets_out: list[dict[str, Any]] = []
    fields_out: list[dict[str, Any]] = []

    seq = 0
    for sname in wb.sheetnames:
        if "instruction" in sname.lower():
            continue
        ws = wb[sname]
        if ws.max_row < 3 or ws.max_column < 2:
            continue

        # Discover metadata-row labels in column 1
        meta_rows: list[tuple[int, str]] = []
        for r in range(1, min(15, ws.max_row + 1)):
            v = ws.cell(r, 1).value
            if v:
                meta_rows.append((r, str(v).strip()))

        # Required header rows: Name (1), Description (2), Data Type (3)
        # Module rows: any subsequent row whose label looks like a module.
        name_row = next((r for r, lbl in meta_rows if lbl.lower() == "name"), 1)
        desc_row = next((r for r, lbl in meta_rows if "description" in lbl.lower()), 2)
        type_row = next(
            (r for r, lbl in meta_rows if "data type" in lbl.lower() or lbl.lower() == "type"), 3
        )
        module_rows = [
            (r, lbl) for r, lbl in meta_rows
            if r > type_row and _looks_like_module_label(lbl)
        ]

        sheet_field_count = 0
        for col in range(2, ws.max_column + 1):
            raw_name = ws.cell(name_row, col).value
            if not raw_name:
                continue
            name = str(raw_name).strip()
            if not name:
                continue

            is_required_marker = name.startswith("*")
            field_name_clean = name.lstrip("*").strip()
            description_val = ws.cell(desc_row, col).value
            description_text = str(description_val).strip() if description_val else None

            dtype_raw = ws.cell(type_row, col).value
            data_type, max_length = _parse_data_type(dtype_raw)

            required_modules: list[str] = []
            for r, lbl in module_rows:
                cell_val = ws.cell(r, col).value
                if cell_val and "required" in str(cell_val).strip().lower():
                    required_modules.append(lbl)

            required = is_required_marker or bool(required_modules)
            seq += 1
            sheet_field_count += 1

            fields_out.append(
                {
                    "field_name": field_name_clean,
                    "display_name": field_name_clean,
                    "description": description_text,
                    "required": required,
                    "data_type": data_type,
                    "max_length": max_length,
                    "format_mask": "YYYY/MM/DD" if data_type.lower() == "date" else None,
                    "sample_value": None,
                    "lookup_type": None,
                    "validation_notes": None,
                    "sequence": seq,
                    "sheet_name": sname,
                    "required_modules": required_modules,
                }
            )

        sheets_out.append({"sheet_name": sname, "sequence": len(sheets_out), "field_count": sheet_field_count})

    return {
        "business_object": business_object,
        "description": description,
        "sheets": sheets_out,
        "fields": fields_out,
    }
